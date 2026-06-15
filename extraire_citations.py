"""
Extraction de citations qualitatives — Affaire Valérie Bacot
=============================================================
Lit le corpus local, extrait les passages les plus représentatifs
de chaque catégorie de narratif, et les ajoute dans l'onglet
"Citations qualitatives" du fichier Excel existant.

Prérequis :
    - analyse_bacot_complet.xlsx  (dans le même dossier)
    - analyse_bacot/resultats_classification.csv
    - corpus_bacot/corpus_bacot.json (ou corpus_final.json)

Installation :
    pip install openpyxl pandas

Usage :
    python extraire_citations.py
"""

import json
import re
import pandas as pd
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

CORPUS_PATHS = [
    Path("corpus_bacot/corpus_final.json"),
    Path("corpus_bacot/corpus_bacot.json"),
]

# Cherche le CSV dans plusieurs emplacements
CSV_PATHS = [
    Path("analyse_bacot/resultats_classification.csv"),
    Path("resultats_classification.csv"),
]

# Cherche l'Excel dans plusieurs emplacements
EXCEL_PATHS = [
    Path("analyse_bacot_complet.xlsx"),
    Path("analyse_bacot/analyse_bacot_complet.xlsx"),
]

NOM_ONGLET          = "💬 Citations qualitatives"
N_CITATIONS_PAR_CAT = 6
LONGUEUR_EXTRAIT    = 500

# ─── Catégories ───────────────────────────────────────────────────────────────

CATEGORIES = [
    'soutien_victime', 'remise_en_question', 'legitime_defense',
    'discours_feministe', 'emprise_psychologique', 'silence_collectif',
    'sensationnalisme', 'jugement_moral',
]

LABELS_FR = {
    'soutien_victime':        '💙 Soutien à la victime',
    'remise_en_question':     '❓ Remise en question',
    'legitime_defense':       '⚖️ Légitime défense',
    'discours_feministe':     '✊ Discours féministe',
    'emprise_psychologique':  '🔗 Emprise psychologique',
    'silence_collectif':      '🤫 Silence collectif',
    'sensationnalisme':       '📺 Sensationnalisme',
    'jugement_moral':         '🔍 Jugement moral',
}

DESCRIPTIONS = {
    'soutien_victime':        "Compassion, solidarité, validation du geste comme acte de survie.",
    'remise_en_question':     "Doute sur les choix, évocation d'alternatives, peut aller jusqu'à la condamnation.",
    'legitime_defense':       "Cadre juridique, légitime défense différée, précédent Jacqueline Sauvage.",
    'discours_feministe':     "Approche systémique — patriarcat, féminicide, continuum des violences.",
    'emprise_psychologique':  "Mécanisme d'emprise, contrôle coercitif, traumatisme, prostitution forcée.",
    'silence_collectif':      "Complicité passive — 'tout le monde savait' — entourage, institutions.",
    'sensationnalisme':       "Fait divers spectaculaire, true crime, sans dimension critique.",
    'jugement_moral':         "Jugement sur la légitimité de la peine, culpabilité morale.",
}

TERMES_PAR_CAT = {
    'soutien_victime': [
        'courage', 'victime', 'survie', 'soutien', 'innocente',
        'libre', 'méritait', 'avait raison', 'bravo', 'admire',
        'aurait fait pareil', 'compréhensible', 'normal',
        'comprends', 'protéger', 'échappatoire',
    ],
    'remise_en_question': [
        'pourquoi', 'partir', 'quitter', 'appeler', 'police',
        'quand même', 'prémédita', 'choix', 'aurait pu', 'devait',
        'autre solution', 'responsable', 'coupable', 'mais',
        'condamn', 'mérite sa peine',
    ],
    'legitime_defense': [
        'légitime défense', 'legitime defense', 'jacqueline sauvage',
        'loi', 'réforme', 'juridique', 'droit', 'code pénal',
        'différée', 'jurisprudence', 'angleterre', 'canada',
        'état de nécessité', 'contrainte',
    ],
    'discours_feministe': [
        'féminicide', 'patriarcat', 'systémique', 'violences faites',
        'féminisme', 'nous toutes', 'metoo', 'structurel',
        'domination', 'genre', 'inégalité', 'femmes',
        'droits', 'oppression',
    ],
    'emprise_psychologique': [
        'emprise', 'contrôle', 'manipulation', 'traumatisme',
        'peur', 'terrorisée', 'prostituée', 'forcée', 'obligée',
        'conditionnée', 'syndrome', 'dépendance', 'isolement',
        'échappatoire', 'survie', 'victime',
    ],
    'silence_collectif': [
        'tout le monde savait', 'savaient', 'silence', 'complice',
        'entourage', 'institution', 'signalement', 'voisins',
        'école', 'médecin', 'fermé les yeux', 'rien fait',
        'savait', 'taire',
    ],
    'sensationnalisme': [
        'choquant', 'horrible', 'incroyable', 'true crime',
        'abonner', 'like', 'regarder', 'story', 'affaire',
        'crime', 'chaîne', 'vidéo', 'documentaire', 'histoire',
    ],
    'jugement_moral': [
        'mérite', 'méritait', 'coupable', 'punition', 'justice',
        'condamner', 'pardon', 'moral', 'responsable',
        'faute', 'sévère', 'clément', 'verdict', 'sentence',
    ],
}

COULEURS_CATS = {
    'soutien_victime':       'DDEEFF',
    'remise_en_question':    'FFD6D6',
    'legitime_defense':      'E8F5E8',
    'discours_feministe':    'F0E8FF',
    'emprise_psychologique': 'FFE8F0',
    'silence_collectif':     'FFE8CC',
    'sensationnalisme':      'FFE8E8',
    'jugement_moral':        'FFF0CC',
}

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)


# ─── Utilitaires ──────────────────────────────────────────────────────────────

def border_fine():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, bg="1F3864", fg="FFFFFF", size=10, bold=True):
    cell.font = Font(bold=bold, color=fg, size=size, name='Calibri')
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_cell(cell, bg=None, bold=False, wrap=True, align='left', size=10):
    cell.font = Font(bold=bold, size=size, name='Calibri')
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical='top', wrap_text=wrap)


def extraire_meilleur_passage(texte: str, categorie: str, n_chars: int = LONGUEUR_EXTRAIT) -> str:
    """
    Extrait le passage le plus pertinent d'un texte pour une catégorie.
    Utilise une fenêtre glissante pour trouver la zone la plus dense
    en termes caractéristiques de la catégorie.
    """
    if not texte or len(texte.strip()) < 30:
        return ""

    termes = TERMES_PAR_CAT.get(categorie, [])
    texte_lower = texte.lower()

    # Si le texte est plus court que la fenêtre, on le prend en entier
    if len(texte) <= n_chars:
        return f"« {texte.strip()} »"

    meilleur_pos   = 0
    meilleur_score = -1

    # Fenêtre glissante
    pas = max(20, n_chars // 15)
    for i in range(0, len(texte) - n_chars, pas):
        fenetre = texte_lower[i:i + n_chars]
        score = sum(fenetre.count(t) for t in termes)
        if score > meilleur_score:
            meilleur_score = score
            meilleur_pos = i

    extrait = texte[meilleur_pos:meilleur_pos + n_chars].strip()

    # Tente de trouver une phrase complète au début
    for punct in ['. ', '.\n', '? ', '! ', '\n\n']:
        idx = extrait.find(punct)
        if 0 < idx < 100:
            extrait = extrait[idx + len(punct):].strip()
            break

    prefix = "... " if meilleur_pos > 100 else ""
    suffix = " ..." if meilleur_pos + n_chars < len(texte) - 100 else ""

    return f"« {prefix}{extrait}{suffix} »"


# ─── Chargement ───────────────────────────────────────────────────────────────

def charger_corpus() -> dict:
    for path in CORPUS_PATHS:
        if path.exists():
            log.info(f"Corpus : {path}")
            with open(path, encoding='utf-8') as f:
                corpus = json.load(f)
            log.info(f"  → {len(corpus)} documents")
            return {doc.get('url', ''): doc for doc in corpus if doc.get('url')}
    log.error("Corpus introuvable.")
    return {}


def charger_resultats() -> pd.DataFrame:
    for path in CSV_PATHS:
        if path.exists():
            log.info(f"CSV classification : {path}")
            return pd.read_csv(path, encoding='utf-8-sig')
    log.error("CSV de classification introuvable.")
    return pd.DataFrame()


def trouver_excel() -> Path | None:
    for path in EXCEL_PATHS:
        if path.exists():
            log.info(f"Excel trouvé : {path}")
            return path
    log.error("Fichier Excel introuvable dans les emplacements suivants :")
    for p in EXCEL_PATHS:
        log.error(f"  {p}")
    return None


# ─── Extraction ───────────────────────────────────────────────────────────────

def extraire_citations(df: pd.DataFrame, corpus_index: dict) -> dict:
    citations_par_cat = {}

    for cat in CATEGORIES:
        log.info(f"Catégorie : {cat}")
        col_score = f"score_{cat}"

        if col_score not in df.columns:
            log.warning(f"  Colonne {col_score} absente du CSV")
            citations_par_cat[cat] = []
            continue

        # Prend les meilleurs documents de cette catégorie
        subset = df[df['categorie_dominante'] == cat].copy()
        subset = subset.sort_values(col_score, ascending=False)

        log.info(f"  {len(subset)} documents dans cette catégorie")

        citations = []

        for _, row in subset.iterrows():
            if len(citations) >= N_CITATIONS_PAR_CAT:
                break

            url  = row.get('url', '')
            doc  = corpus_index.get(url, {})
            texte = doc.get('text', '').strip()

            if not texte:
                log.debug(f"  Pas de texte pour {url[:50]}")
                continue

            nb_mots = len(texte.split())
            if nb_mots < 10:
                log.debug(f"  Texte trop court ({nb_mots} mots)")
                continue

            passage = extraire_meilleur_passage(texte, cat)
            if not passage:
                continue

            type_doc = row.get('type_doc', '')
            if not type_doc:
                source = doc.get('source', '')
                type_doc = 'commentaire' if source == 'youtube_commentaire' else 'article'

            sitename = row.get('sitename', '')
            if not sitename or str(sitename) == 'nan':
                sitename = doc.get('sitename', doc.get('chaine', '?'))

            citations.append({
                'type':     type_doc,
                'source':   str(sitename)[:40],
                'date':     str(row.get('date', ''))[:10],
                'titre':    str(row.get('titre', ''))[:70],
                'score':    int(row.get(col_score, 0)),
                'mots':     nb_mots,
                'citation': passage,
                'url':      url,
            })

        log.info(f"  → {len(citations)} citations extraites")
        citations_par_cat[cat] = citations

    return citations_par_cat


# ─── Écriture Excel ───────────────────────────────────────────────────────────

def remplir_onglet_citations(citations_par_cat: dict, excel_path: Path):
    log.info(f"Ouverture de {excel_path}...")
    wb = load_workbook(excel_path)

    # Supprime l'ancien onglet et en crée un nouveau
    if NOM_ONGLET in wb.sheetnames:
        del wb[NOM_ONGLET]

    # Insère en 3e position
    ws = wb.create_sheet(NOM_ONGLET, 2)

    # En-tête
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "CITATIONS QUALITATIVES PAR CATÉGORIE DE NARRATIF"
    style_header(c, bg="1F3864", size=14)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = (
        f"Top {N_CITATIONS_PAR_CAT} extraits par catégorie, triés par score de pertinence. "
        "Base documentaire pour citations dans l'article analytique."
    )
    style_header(c, bg="2E75B6", size=9, bold=False)
    ws.row_dimensions[2].height = 18

    row_cursor = 3

    for cat in CATEGORIES:
        citations = citations_par_cat.get(cat, [])
        bg_cat    = COULEURS_CATS.get(cat, 'FFFFFF')
        label     = LABELS_FR.get(cat, cat)
        desc      = DESCRIPTIONS.get(cat, '')

        # Séparateur catégorie
        row_cursor += 1
        ws.merge_cells(f'A{row_cursor}:F{row_cursor}')
        c = ws.cell(row=row_cursor, column=1)
        c.value = f"{label}  —  {desc}"
        style_header(c, bg="2E75B6", size=10)
        ws.row_dimensions[row_cursor].height = 32
        row_cursor += 1

        if not citations:
            ws.merge_cells(f'A{row_cursor}:F{row_cursor}')
            c = ws.cell(row=row_cursor, column=1)
            c.value = "Aucun document avec texte accessible pour cette catégorie."
            style_cell(c, bg='F5F5F5', size=9)
            ws.row_dimensions[row_cursor].height = 18
            row_cursor += 1
            continue

        # En-têtes colonnes
        headers = ['#', 'Type', 'Source', 'Date', 'Score', 'Extrait représentatif']
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=row_cursor, column=j)
            c.value = h
            style_header(c, bg="1F3864", size=9)
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        # Citations
        for k, cit in enumerate(citations, 1):
            type_emoji = "📰" if cit['type'] == 'article' else "💬"

            vals = [
                k,
                f"{type_emoji} {cit['type'].capitalize()}",
                cit['source'],
                cit['date'],
                cit['score'],
                cit['citation'],
            ]

            for j, val in enumerate(vals, 1):
                c = ws.cell(row=row_cursor, column=j)
                c.value = val

                if j == 6:
                    c.font      = Font(size=10, name='Calibri', italic=True)
                    c.fill      = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    style_cell(c, bg=bg_cat, size=9,
                               align='center' if j in [1, 4, 5] else 'left')

                c.border = border_fine()

            ws.row_dimensions[row_cursor].height = 100
            row_cursor += 1

        ws.row_dimensions[row_cursor].height = 8
        row_cursor += 1

    # Largeurs
    for i, w in enumerate([4, 14, 28, 11, 8, 85], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'

    wb.save(excel_path)
    log.info(f"✓ Onglet mis à jour dans {excel_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def run():
    log.info("=" * 60)
    log.info("EXTRACTION DE CITATIONS QUALITATIVES")
    log.info("=" * 60)

    corpus_index = charger_corpus()
    if not corpus_index:
        return

    df = charger_resultats()
    if df.empty:
        return

    excel_path = trouver_excel()
    if not excel_path:
        return

    log.info("=" * 60)
    log.info("Extraction des passages...")
    log.info("=" * 60)

    citations_par_cat = extraire_citations(df, corpus_index)

    total = sum(len(v) for v in citations_par_cat.values())
    log.info(f"\nTotal citations : {total}")
    for cat, cits in citations_par_cat.items():
        status = "✓" if cits else "✗"
        log.info(f"  {status} {cat:<35} : {len(cits)} citations")

    log.info("=" * 60)
    log.info("Mise à jour Excel...")
    log.info("=" * 60)

    remplir_onglet_citations(citations_par_cat, excel_path)

    log.info(f"""
╔══════════════════════════════════════════════════╗
║      CITATIONS QUALITATIVES EXTRAITES ✓          ║
╠══════════════════════════════════════════════════╣
║  Citations totales  : {total:>5}                   ║
║  Catégories         :     8                   ║
╚══════════════════════════════════════════════════╝
→ Ouvre : {excel_path}
→ Onglet : {NOM_ONGLET}
    """)


if __name__ == "__main__":
    run()
